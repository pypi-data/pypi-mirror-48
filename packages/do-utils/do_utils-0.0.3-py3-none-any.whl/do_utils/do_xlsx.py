# -*- coding: utf-8 -*-
# -*- author: Jiangtao -*-

"""Do xlsx write

"""


import os
import logging

from xlwt import Workbook
from openpyxl import Workbook as LaWorkbook


class XlsxHandler(object):

    def __init__(self, save_name):
        logging.info('satrt deal with xlsx')
        self.wb = Workbook(encoding='utf-8')
        self.save_name = save_name
        pass

    def __del__(self):
        logging.info('complete, saving xlsx...')
        if '.' in self.save_name:
            self.save_name = self.save_name.split('.')[0]
        save_path = os.path.join(os.path.dirname(__name__), 'xlsx_out')
        if not os.path.exists(save_path):
            os.mkdir(save_path)
        save_name = '{}.xls'.format(self.save_name)
        save_name = os.path.join(save_path, save_name)
        self.wb.save(save_name)
        logging.info('saved xlsx to %s', save_name)

    def write(self, title_list, data_list, sheet_name=None):
        """
        write to xlsx with title & data
        :param title_list: list or tuple, title
        :param data_list: list or tuple, data
        :param sheet_name: str, name of sheet
        :return: bool
        """
        if not all([title_list, isinstance(title_list, (list, tuple))]):
            return False
        if not all([data_list, isinstance(data_list, (list, tuple))]):
            return False

        # wb = Workbook(encoding='utf-8')

        ws = self.wb.add_sheet(sheet_name or '1')

        title = tuple(title_list)
        data = tuple(data_list)

        for i, t in enumerate(title):
            ws.write(0, i, t)

        for i, d in enumerate(data):
            for idx in range(len(title)):
                ws.write(i+1, idx, d[idx])

        # if '.' in file_name:
        #     file_name = file_name.split('.')[0]
        # wb.save('{}.xlsx'.format(file_name))

        return True


class LargeXlsxHandler(object):

    def __init__(self, save_name):
        logging.info('satrt deal with large xlsx')
        self.wb = LaWorkbook()
        self.save_name = save_name
        pass

    def __del__(self):
        logging.info('complete, saving large xlsx...')
        if '.' in self.save_name:
            self.save_name = self.save_name.split('.')[0]
        save_path = os.path.join(os.path.dirname(__name__), 'xlsx_out')
        if not os.path.exists(save_path):
            os.mkdir(save_path)
        save_name = '{}.xls'.format(self.save_name)
        save_name = os.path.join(os.path.dirname(__name__), save_name)
        self.wb.save(save_name)
        logging.info('saved large xlsx to %s', save_name)

    def write(self, title_list, data_list, sheet_name=None, new_sheet=False):
        """
        write to xlsx with title & data
        :param title_list: list or tuple, title
        :param data_list: list or tuple, data
        :param sheet_name: str, name of sheet
        :return: bool
        """
        if not all([title_list, isinstance(title_list, (list, tuple))]):
            return False
        if not all([data_list, isinstance(data_list, (list, tuple))]):
            return False

        # wb = Workbook(encoding='utf-8')

        if not new_sheet:
            ws = self.wb.active
        else:
            ws = self.wb.create_sheet(sheet_name)

        title = list(title_list)
        data = list(data_list)

        # for i, t in enumerate(title):
        #     ws.write(0, i, t)
        ws.append(title)

        # for i, d in enumerate(data):
        #     for idx in range(len(title)):
        #         ws.write(i+1, idx, d[idx])
        for d in data:
            ws.append(d)

        # if '.' in file_name:
        #     file_name = file_name.split('.')[0]
        # wb.save('{}.xlsx'.format(file_name))

        return True


__all__ = (XlsxHandler, LargeXlsxHandler)


def main():
    file_name = 'test'
    title = ['id', 'name']
    data = [(123, 'xiaomu'), (124, 'xiaoming')]
    sheet_name = 'name'
    # do_xlsx = XlsxHandler(save_name=file_name)
    # do_xlsx.write(title_list=title, data_list=data, sheet_name=sheet_name)

    do_laxlsx = LargeXlsxHandler(save_name=file_name)
    do_laxlsx.write(title_list=title, data_list=data, sheet_name=sheet_name, new_sheet=True)


if __name__ == '__main__':
    main()
