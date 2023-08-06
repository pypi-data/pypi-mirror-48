from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from typing import List, Optional, Type
from enum import Enum


class ExcelMethods(str, Enum):
    READ = 'r'
    WRITE = 'w'


class KaWorkBook(Workbook):
    def write_data(self, datas: List[List[str]],
                   sheetname: Optional[str] = None) -> None:
        if sheetname:
            sh = self.create_sheet(sheetname)
        else:
            sh = self.active
        for data in datas:
            sh.append(data)


class OpenExcel(object):
    def __init__(self, filename: str, methods: Type[ExcelMethods]):
        self.filename = filename
        self.methods = methods
        self.cursor = self.enter_methods()

    def write_cursor(self):
        wb = KaWorkBook()
        return wb

    def read_cursor(self):
        wb = load_workbook(self.filename)
        return wb

    def enter_methods(self):
        d = {
            ExcelMethods.WRITE: self.write_cursor,
            ExcelMethods.READ: self.read_cursor,
        }
        if self.methods not in d.keys():
            raise KeyboardInterrupt
        return d.get(self.methods, None)()

    def exit_methods(self):
        if self.methods == ExcelMethods.WRITE:
            return self.cursor.save(self.filename)
        elif self.methods == ExcelMethods.READ:
            return self.cursor.close()
        else:
            return None

    def __enter__(self):
        return self.cursor

    def __exit__(self, exc_type, exc_value, exc_tb):
        self.exit_methods()
