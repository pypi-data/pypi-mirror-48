from django.apps import apps
from openpyxl import load_workbook
from .models import Subdivision


def apply_update(data):
    models = apps.get_app_config("glossary").get_models()
    models = sorted(models, key=lambda m: m.get_foreign_key_count())
    for model in models:
        records = data[model._meta.model_name]
        for record in records:
            uuid = record["uuid"]
            obj_data = model.clean_dump_data(record)

            try:
                model.all_objects.get(uuid=uuid)
                model.all_objects.filter(uuid=uuid).update(**obj_data)
            except model.DoesNotExist:
                model.all_objects.create(**obj_data)


def parse_subdivisions_excel(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    i = 2
    while ws['H' + str(i)].value:
        print('{} - {}'.format(str(i), ws['H' + str(i)].value))
        Subdivision.objects.get_or_create(name=ws['H' + str(i)].value)
        i += 1
